"""Tests for the ABCD report exporters and their routes.

Covers the pure CSV/XLSX builders in ``app.exports`` and the run-level and
job-level export endpoints, including the empty-job and bad-format guards.
Uses ``metadata.create_all`` directly (tests own their schema).
"""

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine

from app import exports
from app.deps import get_reader, get_writer
from app.main import app
from src.db.reader import ClipScribeReaderDB
from src.db.schema import metadata_obj, parser_results_table, runs_table
from src.db.writer import ClipScribeWriterDB

PARENT = "job-parent"
CHILD1 = "job-child-1"
CHILD2 = "job-child-2"
RUN1 = "run-1"
RUN2 = "run-2"


def _rows(passed_first: bool) -> list[dict]:
    return [
        {
            "feature_category": "Attract",
            "feature_name": "Dynamic Start",
            "feature_criteria": "hooks in first 5s",
            "evaluation": passed_first,
            "llm_explanation": "fast cuts early",
            "llm_prompt": "prompt A",
        },
        {
            "feature_category": "Brand",
            "feature_name": "Brand Mention",
            "feature_criteria": "brand named",
            "evaluation": False,
            "llm_explanation": "no brand",
            "llm_prompt": "prompt B",
        },
    ]


# ── pure builders ────────────────────────────────────────────────────────────


def test_run_csv_header_and_verdicts():
    out = exports.run_csv(_rows(True)).decode()
    lines = out.splitlines()
    assert lines[0] == "Category,Feature,Criteria,Result,Explanation,Prompt"
    assert "Attract,Dynamic Start" in lines[1]
    assert lines[1].split(",")[3] == "PASS"
    assert lines[2].split(",")[3] == "FAIL"


def test_run_xlsx_has_detail_and_scores_sheets():
    wb = openpyxl.load_workbook(io.BytesIO(exports.run_xlsx("ad.mp4", _rows(True))))
    assert wb.sheetnames == ["Detail", "Scores"]
    scores = wb["Scores"]
    # Header + Attract + Brand + Total.
    values = {row[0].value: (row[1].value, row[2].value) for row in scores.iter_rows()}
    assert values["Total"] == (1, 2)


def test_job_csv_prefixes_video_column():
    reports = [
        exports.RunReport(RUN1, "first.mp4", _rows(True)),
        exports.RunReport(RUN2, "second.mp4", _rows(False)),
    ]
    lines = exports.job_csv(reports).decode().splitlines()
    assert lines[0].startswith("Video,Category")
    assert lines[1].startswith("first.mp4,")
    assert any(line.startswith("second.mp4,") for line in lines)


def test_job_xlsx_summary_and_unique_sheets_per_run():
    reports = [
        exports.RunReport(RUN1, "same name.mp4", _rows(True)),
        exports.RunReport(RUN2, "same name.mp4", _rows(False)),
    ]
    wb = openpyxl.load_workbook(io.BytesIO(exports.job_xlsx(reports)))
    assert wb.sheetnames[0] == "Summary"
    # Two detail sheets, uniquely titled despite identical video names.
    assert len(wb.sheetnames) == 3
    assert len(set(wb.sheetnames)) == 3
    summary = wb["Summary"]
    totals = [row for row in summary.iter_rows(values_only=True) if row[0] == "Total"]
    assert totals[0][1:3] == (1, 4)  # 1 passed of 4 criteria across both runs


def test_export_filename_slugs_unsafe_chars():
    assert exports.export_filename("ad: name/v2_abcd", "xlsx") == "ad_name_v2_abcd.xlsx"
    assert exports.export_filename("", "csv") == "export.csv"


# ── routes ───────────────────────────────────────────────────────────────────


@pytest.fixture
def ctx(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path / 'exp.db'}")
    metadata_obj.create_all(engine)
    reader = ClipScribeReaderDB(engine=engine)
    writer = ClipScribeWriterDB(engine=engine)

    with engine.begin() as conn:
        conn.execute(
            runs_table.insert(),
            [
                {"run_id": RUN1, "video_name": "first.mp4"},
                {"run_id": RUN2, "video_name": "second.mp4"},
            ],
        )
        for run_id, rows in ((RUN1, _rows(True)), (RUN2, _rows(False))):
            conn.execute(
                parser_results_table.insert(),
                [{"run_id": run_id, "platform": "youtube", **r} for r in rows],
            )

    writer.create_job(job_id=PARENT, mode="full", video_name="batch")
    writer.create_job(
        job_id=CHILD1,
        mode="full",
        parent_job_id=PARENT,
        run_id=RUN1,
        video_name="first.mp4",
        status="completed",
    )
    writer.create_job(
        job_id=CHILD2,
        mode="full",
        parent_job_id=PARENT,
        run_id=RUN2,
        video_name="second.mp4",
        status="completed",
    )

    app.dependency_overrides[get_reader] = lambda: reader
    app.dependency_overrides[get_writer] = lambda: writer
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.clear()


def test_run_export_xlsx_route(ctx):
    resp = ctx.get(f"/runs/{RUN1}/parser/export?format=xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == exports.CONTENT_TYPES["xlsx"]
    assert "attachment" in resp.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Detail", "Scores"]


def test_run_export_csv_route(ctx):
    resp = ctx.get(f"/runs/{RUN1}/parser/export?format=csv")
    assert resp.status_code == 200
    assert resp.text.splitlines()[0].startswith("Category,")


def test_job_export_xlsx_route(ctx):
    resp = ctx.get(f"/jobs/{PARENT}/export?format=xlsx")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames[0] == "Summary"
    assert len(wb.sheetnames) == 3  # Summary + one per run


def test_job_export_csv_stacks_runs(ctx):
    resp = ctx.get(f"/jobs/{PARENT}/export?format=csv")
    assert resp.status_code == 200
    body = resp.text
    assert "first.mp4," in body
    assert "second.mp4," in body


def test_export_bad_format_400(ctx):
    assert ctx.get(f"/runs/{RUN1}/parser/export?format=pdf").status_code == 400
    assert ctx.get(f"/jobs/{PARENT}/export?format=pdf").status_code == 400


def test_export_unknown_404(ctx):
    assert ctx.get("/runs/nope/parser/export").status_code == 404
    assert ctx.get("/jobs/nope/export").status_code == 404


def test_job_export_no_completed_runs_409(ctx, tmp_path):
    # A fresh parent whose only child is still running → nothing to export.
    engine = create_engine(f"sqlite:///{tmp_path / 'empty.db'}")
    metadata_obj.create_all(engine)
    reader = ClipScribeReaderDB(engine=engine)
    writer = ClipScribeWriterDB(engine=engine)
    writer.create_job(job_id="p2", mode="full", video_name="batch")
    writer.create_job(job_id="c2", mode="full", parent_job_id="p2", status="running")
    app.dependency_overrides[get_reader] = lambda: reader
    assert ctx.get("/jobs/p2/export").status_code == 409
