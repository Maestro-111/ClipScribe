"""ABCD report exporters (web-app-plan §6 extension).

Builds CSV / XLSX bytes from persisted ``parser_results`` rows so the frontend
can download a run's or a whole job's evaluation as a spreadsheet. Platform
agnostic: it reads whatever columns the reader returns (mirrors
``writer.save_parser_results``), so non-YouTube platforms export the common
fields too.

Pure over already-fetched data — routes do the DB reads and hand the rows in,
which keeps these functions unit-testable without a database and keeps the
import light (no torch, no reader). The parser's own ``YouTubeReportWriter``
stays untouched: it is worker-side, file-path-based, and YouTube-specific,
whereas this serves on-demand HTTP downloads from the DB.
"""

from __future__ import annotations

import csv
import io
import re
from collections import Counter
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Category display order for the scores summary; unknown categories are appended
# in first-seen order after these.
CATEGORY_ORDER = ["Attract", "Brand", "Connect", "Direct"]

# Per-criterion detail columns, in export order. ``llm_prompt`` is last (widest,
# least scanned) so a reader sees the verdict + reasoning first.
DETAIL_FIELDS: list[tuple[str, str]] = [
    ("feature_category", "Category"),
    ("feature_name", "Feature"),
    ("feature_criteria", "Criteria"),
    ("evaluation", "Result"),
    ("llm_explanation", "Explanation"),
    ("llm_prompt", "Prompt"),
]

CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


@dataclass(frozen=True)
class RunReport:
    """One run's data for a job-level export."""

    run_id: str
    video_name: str
    rows: list[dict]


VALID_FORMATS = frozenset({"csv", "xlsx"})


def export_filename(stem: str, fmt: str) -> str:
    """A safe download filename: slugged stem + format extension."""
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._") or "export"
    return f"{safe}.{fmt}"


def _result_text(evaluation: object) -> str:
    """Render the boolean verdict as PASS / FAIL (blank if unknown)."""
    if evaluation is None:
        return ""
    return "PASS" if evaluation else "FAIL"


def _detail_value(row: dict, field: str) -> str:
    if field == "evaluation":
        return _result_text(row.get("evaluation"))
    value = row.get(field)
    return "" if value is None else str(value)


def _category_order(rows: list[dict]) -> list[str]:
    """Known categories first (fixed order), then any extras in first-seen order."""
    seen = [str(r.get("feature_category") or "") for r in rows]
    extras = [c for c in dict.fromkeys(seen) if c and c not in CATEGORY_ORDER]
    return CATEGORY_ORDER + extras


def _scores(rows: list[dict]) -> tuple[list[tuple[str, int, int]], int, int]:
    """Per-category (category, passed, total) plus overall (passed, total)."""
    total: Counter = Counter()
    passed: Counter = Counter()
    for r in rows:
        cat = str(r.get("feature_category") or "")
        total[cat] += 1
        if r.get("evaluation"):
            passed[cat] += 1
    per_category = [
        (cat, passed.get(cat, 0), total.get(cat, 0))
        for cat in _category_order(rows)
        if total.get(cat, 0)
    ]
    return per_category, sum(passed.values()), sum(total.values())


def _score_pct(passed: int, total: int) -> float:
    return round(passed / total, 4) if total else 0.0


# ── CSV ─────────────────────────────────────────────────────────────────────


def run_csv(rows: list[dict]) -> bytes:
    """One run's criteria as a flat CSV table."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _key, label in DETAIL_FIELDS])
    for r in rows:
        writer.writerow([_detail_value(r, key) for key, _label in DETAIL_FIELDS])
    return buf.getvalue().encode("utf-8")


def job_csv(reports: list[RunReport]) -> bytes:
    """All runs' criteria stacked into one CSV, prefixed with a Video column.

    CSV has no sheets, so the job export collapses to a single table; the Video
    column is what lets a reader split it back out (or pivot in Excel).
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Video", *(label for _key, label in DETAIL_FIELDS)])
    for report in reports:
        for r in report.rows:
            writer.writerow(
                [report.video_name, *(_detail_value(r, k) for k, _l in DETAIL_FIELDS)]
            )
    return buf.getvalue().encode("utf-8")


# ── XLSX ────────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True)


def _autosize(ws: Worksheet, max_width: int = 80) -> None:
    """Rough column autosize from cell text length, capped so prose stays sane."""
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max(len(line) for line in str(cell.value).splitlines() or [""])
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, max_width)


def _write_detail_sheet(ws: Worksheet, rows: list[dict]) -> None:
    ws.append([label for _key, label in DETAIL_FIELDS])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for r in rows:
        ws.append([_detail_value(r, key) for key, _label in DETAIL_FIELDS])
    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_scores_sheet(ws: Worksheet, rows: list[dict]) -> None:
    per_category, total_passed, total_all = _scores(rows)
    ws.append(["Category", "Passed", "Total", "Score"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for cat, passed, total in per_category:
        ws.append([cat, passed, total, _score_pct(passed, total)])
    total_row = ["Total", total_passed, total_all, _score_pct(total_passed, total_all)]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = "0.0%"
    _autosize(ws)


def _safe_sheet_title(name: str, index: int, used: set[str]) -> str:
    """A unique, Excel-legal (<=31 char, no ``[]:*?/\\``) sheet title.

    Prefixed with the 1-based run number so order is obvious and collisions
    between same-named videos are impossible.
    """
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "run"
    prefix = f"{index}. "
    title = (prefix + cleaned)[:31].strip()
    base = title
    n = 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def run_xlsx(video_name: str, rows: list[dict]) -> bytes:
    """One run: a Detail sheet (per-criterion) plus a Scores summary sheet."""
    wb = Workbook()
    detail = wb.active
    detail.title = "Detail"
    _write_detail_sheet(detail, rows)
    _write_scores_sheet(wb.create_sheet("Scores"), rows)
    return _to_bytes(wb)


def job_xlsx(reports: list[RunReport]) -> bytes:
    """Whole job: a Summary sheet across runs, then one Detail sheet per run."""
    wb = Workbook()
    _write_job_summary(wb.active, reports)
    used: set[str] = set()
    for i, report in enumerate(reports, start=1):
        title = _safe_sheet_title(report.video_name, i, used)
        _write_detail_sheet(wb.create_sheet(title), report.rows)
    return _to_bytes(wb)


def _write_job_summary(ws: Worksheet, reports: list[RunReport]) -> None:
    """One row per run (video, passed/total, pass rate) + a totals row."""
    ws.title = "Summary"
    ws.append(["Video", "Passed", "Total", "Pass rate"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    grand_passed = grand_total = 0
    for report in reports:
        _per_cat, passed, total = _scores(report.rows)
        grand_passed += passed
        grand_total += total
        ws.append([report.video_name, passed, total, _score_pct(passed, total)])
    ws.append(
        ["Total", grand_passed, grand_total, _score_pct(grand_passed, grand_total)]
    )
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
    # Show the rate column as a percentage.
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = "0.0%"
    ws.freeze_panes = "A2"
    _autosize(ws)


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
